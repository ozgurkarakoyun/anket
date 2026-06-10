# -*- coding: utf-8 -*-
"""
FJS-12 (Forgotten Joint Score) Web Application
Author: Doç. Dr. Özgür Karakoyun project
Stack: Flask + SQLite, deployable on Railway
Procedures: knee, hip, osseointegration, socket_prosthesis
"""

import os
import json
import csv
import io
import sqlite3
from datetime import datetime
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, abort, Response, flash, g
)

from openpyxl import Workbook as ExcelWorkbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from translations import LANGUAGES, QUESTIONS, ANSWERS, T, t, PROCEDURES, AMPUTEE_PROCEDURES
from scoring import calculate_fjs_score, score_label

DB_PATH = os.environ.get('DB_PATH', os.path.join(os.path.dirname(__file__), 'fjs.db'))
SECRET_KEY = os.environ.get('SECRET_KEY', 'change-me-dev-only-secret-key-12345')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin1234')
DEFAULT_LANG = 'tr'

AMPUTATION_LEVEL_LABELS = {
    'transfemoral': 'Transfemoral (Diz üstü amputasyon)',
    'knee_disart': 'Diz dezartikülasyonu',
    'transtibial': 'Transtibial (Diz altı amputasyon)',
    'hip_disart': 'Kalça dezartikülasyonu',
}

app = Flask(__name__)
app.secret_key = SECRET_KEY

SCHEMA = """
CREATE TABLE IF NOT EXISTS responses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at TEXT DEFAULT (datetime('now')),
    language TEXT,
    name TEXT NOT NULL,
    age INTEGER,
    procedure_type TEXT NOT NULL,
    side TEXT NOT NULL,
    surgery_month_right INTEGER,
    surgery_year_right INTEGER,
    surgery_month_left INTEGER,
    surgery_year_left INTEGER,
    amputation_month_right INTEGER,
    amputation_year_right INTEGER,
    amputation_level_right TEXT,
    amputation_month_left INTEGER,
    amputation_year_left INTEGER,
    amputation_level_left TEXT,
    fjs_score_right REAL,
    fjs_score_left REAL,
    fjs_score_unilateral REAL,
    answers_json TEXT,
    is_archived INTEGER DEFAULT 0,
    archived_at TEXT
);
"""


def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    cursor = conn.execute("PRAGMA table_info(responses)")
    columns = {row[1] for row in cursor.fetchall()}
    for new_col, sql_type in [
        ('amputation_month_right', 'INTEGER'),
        ('amputation_month_left', 'INTEGER'),
        ('answers_json', 'TEXT'),
        ('is_archived', 'INTEGER DEFAULT 0'),
        ('archived_at', 'TEXT'),
    ]:
        if new_col not in columns:
            conn.execute(f"ALTER TABLE responses ADD COLUMN {new_col} {sql_type}")
    conn.commit()
    conn.close()


def get_lang():
    lang = (request.args.get('lang') or request.form.get('lang') or session.get('lang') or DEFAULT_LANG)
    if lang not in LANGUAGES:
        lang = DEFAULT_LANG
    session['lang'] = lang
    return lang


@app.context_processor
def inject_globals():
    lang = session.get('lang', DEFAULT_LANG)
    return {
        'lang': lang,
        'lang_dir': LANGUAGES[lang]['dir'],
        'languages': LANGUAGES,
        't': lambda key, **kw: t(lang, key, **kw),
        'T': T.get(lang, T[DEFAULT_LANG]),
    }


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('is_admin'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return wrapper


def safe_json_loads(value):
    if not value:
        return {}
    try:
        return json.loads(value)
    except (TypeError, ValueError):
        return {}


def answer_text(lang, value):
    if value is None or value == '':
        return ''
    try:
        value = int(value)
    except (TypeError, ValueError):
        return ''
    options = ANSWERS.get(lang) or ANSWERS[DEFAULT_LANG]
    if 0 <= value < len(options):
        return options[value]
    return ''


def amputation_level_text(value):
    if not value:
        return ''
    return AMPUTATION_LEVEL_LABELS.get(value, value)


def parse_int_or_none(value):
    if value in (None, ''):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def active_filter():
    return 'COALESCE(is_archived, 0) = 0'


def get_answer_lists(row):
    payload = safe_json_loads(row['answers_json'])
    side = row['side']
    right_answers = [''] * 12
    left_answers = [''] * 12
    unilateral_answers = [''] * 12

    if side == 'bilateral':
        right = payload.get('right') or []
        left = payload.get('left') or []
        for i in range(12):
            right_answers[i] = right[i] if i < len(right) else ''
            left_answers[i] = left[i] if i < len(left) else ''
    else:
        answers = payload.get('answers') or []
        for i in range(12):
            unilateral_answers[i] = answers[i] if i < len(answers) else ''
        if side == 'right':
            right_answers = unilateral_answers[:]
        elif side == 'left':
            left_answers = unilateral_answers[:]
    return right_answers, left_answers, unilateral_answers


def build_answer_rows(questions, answers, lang):
    rows = []
    for i, question in enumerate(questions):
        value = answers[i] if i < len(answers) else ''
        rows.append({'number': i + 1, 'question': question, 'value': value, 'answer': answer_text(lang, value)})
    return rows


def normalize_scores_for_side(row, side):
    if side == 'bilateral':
        return row['fjs_score_right'], row['fjs_score_left'], None
    single_score = row['fjs_score_unilateral']
    if single_score is None:
        single_score = row['fjs_score_right'] if row['fjs_score_right'] is not None else row['fjs_score_left']
    if side == 'right':
        return single_score, None, single_score
    return None, single_score, single_score


def build_export_headers():
    base_headers = [
        'id', 'created_at', 'language', 'name', 'age', 'procedure_type', 'side',
        'surgery_month_right', 'surgery_year_right', 'surgery_month_left', 'surgery_year_left',
        'amputation_month_right', 'amputation_year_right', 'amputation_level_right', 'amputation_level_right_label',
        'amputation_month_left', 'amputation_year_left', 'amputation_level_left', 'amputation_level_left_label',
        'fjs_right', 'fjs_left', 'fjs_unilateral',
    ]
    answer_headers = []
    for i in range(1, 13):
        answer_headers.extend([f'right_q{i}_value', f'right_q{i}_answer'])
    for i in range(1, 13):
        answer_headers.extend([f'left_q{i}_value', f'left_q{i}_answer'])
    for i in range(1, 13):
        answer_headers.extend([f'unilateral_q{i}_value', f'unilateral_q{i}_answer'])
    return base_headers + answer_headers + ['answers_json']


def build_export_row(row):
    lang = row['language'] if row['language'] in ANSWERS else DEFAULT_LANG
    right_answers, left_answers, unilateral_answers = get_answer_lists(row)
    answer_values = []
    for value in right_answers:
        answer_values.extend([value, answer_text(lang, value)])
    for value in left_answers:
        answer_values.extend([value, answer_text(lang, value)])
    for value in unilateral_answers:
        answer_values.extend([value, answer_text(lang, value)])
    return [
        row['id'], row['created_at'], row['language'], row['name'], row['age'],
        row['procedure_type'], row['side'], row['surgery_month_right'], row['surgery_year_right'],
        row['surgery_month_left'], row['surgery_year_left'], row['amputation_month_right'],
        row['amputation_year_right'], row['amputation_level_right'], amputation_level_text(row['amputation_level_right']),
        row['amputation_month_left'], row['amputation_year_left'], row['amputation_level_left'],
        amputation_level_text(row['amputation_level_left']), row['fjs_score_right'], row['fjs_score_left'],
        row['fjs_score_unilateral'],
    ] + answer_values + [row['answers_json'] or '']


def get_export_table():
    db = get_db()
    rows = db.execute(f'SELECT * FROM responses WHERE {active_filter()} ORDER BY created_at DESC').fetchall()
    return build_export_headers(), [build_export_row(row) for row in rows]


def export_filename(extension):
    return f'fjs_responses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{extension}'


@app.route('/')
def index():
    get_lang()
    return render_template('index.html')


@app.route('/start', methods=['GET', 'POST'])
def patient_info():
    lang = get_lang()
    if request.method == 'POST':
        form = request.form
        name = (form.get('name') or '').strip()
        age = form.get('age', type=int)
        procedure_type = form.get('procedure_type')
        side = form.get('side')
        errors = []
        if not name:
            errors.append('name')
        if not procedure_type or procedure_type not in PROCEDURES:
            errors.append('procedure_type')
        if not side or side not in ('right', 'left', 'bilateral'):
            errors.append('side')

        sm_r = form.get('surgery_month_right', type=int)
        sy_r = form.get('surgery_year_right', type=int)
        sm_l = form.get('surgery_month_left', type=int)
        sy_l = form.get('surgery_year_left', type=int)
        amp_m_r = form.get('amputation_month_right', type=int)
        amp_y_r = form.get('amputation_year_right', type=int)
        amp_l_r = form.get('amputation_level_right')
        amp_m_l = form.get('amputation_month_left', type=int)
        amp_y_l = form.get('amputation_year_left', type=int)
        amp_l_l = form.get('amputation_level_left')

        if procedure_type in ('knee', 'hip', 'osseointegration'):
            if side in ('right', 'bilateral') and not (sm_r and sy_r):
                errors.append('surgery_date_right')
            if side in ('left', 'bilateral') and not (sm_l and sy_l):
                errors.append('surgery_date_left')
        if procedure_type == 'osseointegration':
            if side in ('right', 'bilateral') and not (amp_y_r and amp_l_r):
                errors.append('amputation_right')
            if side in ('left', 'bilateral') and not (amp_y_l and amp_l_l):
                errors.append('amputation_left')
        if procedure_type == 'socket_prosthesis':
            if side in ('right', 'bilateral') and not (amp_m_r and amp_y_r and amp_l_r):
                errors.append('amputation_right')
            if side in ('left', 'bilateral') and not (amp_m_l and amp_y_l and amp_l_l):
                errors.append('amputation_left')
        if errors:
            flash(t(lang, 'required_error'), 'error')
            return render_template('patient_info.html', form=form, errors=errors)

        is_amputee = procedure_type in AMPUTEE_PROCEDURES
        is_socket = procedure_type == 'socket_prosthesis'
        session['patient'] = {
            'language': lang, 'name': name, 'age': age, 'procedure_type': procedure_type, 'side': side,
            'surgery_month_right': sm_r if not is_socket else None,
            'surgery_year_right': sy_r if not is_socket else None,
            'surgery_month_left': sm_l if not is_socket else None,
            'surgery_year_left': sy_l if not is_socket else None,
            'amputation_month_right': amp_m_r if is_socket else None,
            'amputation_year_right': amp_y_r if is_amputee else None,
            'amputation_level_right': amp_l_r if is_amputee else None,
            'amputation_month_left': amp_m_l if is_socket else None,
            'amputation_year_left': amp_y_l if is_amputee else None,
            'amputation_level_left': amp_l_l if is_amputee else None,
        }
        return redirect(url_for('questions'))
    return render_template('patient_info.html', form={}, errors=[])


@app.route('/questions', methods=['GET', 'POST'])
def questions():
    lang = get_lang()
    patient = session.get('patient')
    if not patient:
        return redirect(url_for('patient_info'))
    questions_text = QUESTIONS[lang]
    answer_options = ANSWERS[lang]
    bilateral = patient['side'] == 'bilateral'

    if request.method == 'POST':
        if bilateral:
            answers_right, answers_left = [], []
            for i in range(12):
                answers_right.append(request.form.get(f'q{i}_right', type=int))
                answers_left.append(request.form.get(f'q{i}_left', type=int))
            score_right = calculate_fjs_score(answers_right)
            score_left = calculate_fjs_score(answers_left)
            score_unilateral = None
            answers_payload = {'right': answers_right, 'left': answers_left}
        else:
            ans = [request.form.get(f'q{i}', type=int) for i in range(12)]
            single_score = calculate_fjs_score(ans)
            if patient['side'] == 'right':
                score_right, score_left, score_unilateral = single_score, None, single_score
            else:
                score_right, score_left, score_unilateral = None, single_score, single_score
            answers_payload = {'side': patient['side'], 'answers': ans}

        db = get_db()
        cur = db.execute(
            """INSERT INTO responses
               (language, name, age, procedure_type, side,
                surgery_month_right, surgery_year_right, surgery_month_left, surgery_year_left,
                amputation_month_right, amputation_year_right, amputation_level_right,
                amputation_month_left, amputation_year_left, amputation_level_left,
                fjs_score_right, fjs_score_left, fjs_score_unilateral, answers_json)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                patient['language'], patient['name'], patient['age'], patient['procedure_type'], patient['side'],
                patient['surgery_month_right'], patient['surgery_year_right'], patient['surgery_month_left'], patient['surgery_year_left'],
                patient['amputation_month_right'], patient['amputation_year_right'], patient['amputation_level_right'],
                patient['amputation_month_left'], patient['amputation_year_left'], patient['amputation_level_left'],
                score_right, score_left, score_unilateral, json.dumps(answers_payload, ensure_ascii=False),
            )
        )
        db.commit()
        new_id = cur.lastrowid
        session.pop('patient', None)
        return redirect(url_for('result', response_id=new_id))
    return render_template('questions.html', patient=patient, questions=questions_text, answers=answer_options, bilateral=bilateral)


@app.route('/result/<int:response_id>')
def result(response_id):
    lang = get_lang()
    db = get_db()
    row = db.execute(f'SELECT * FROM responses WHERE id = ? AND {active_filter()}', (response_id,)).fetchone()
    if not row:
        abort(404)
    row_lang = row['language'] if row['language'] in QUESTIONS else DEFAULT_LANG
    questions_for_row = QUESTIONS.get(row_lang, QUESTIONS[DEFAULT_LANG])
    right_answers, left_answers, unilateral_answers = get_answer_lists(row)
    has_answer_details = any(value != '' for value in (right_answers + left_answers + unilateral_answers))
    return render_template(
        'results.html',
        r=row,
        score_label_right=score_label(row['fjs_score_right'], lang) if row['fjs_score_right'] is not None else '',
        score_label_left=score_label(row['fjs_score_left'], lang) if row['fjs_score_left'] is not None else '',
        score_label_uni=score_label(row['fjs_score_unilateral'], lang) if row['fjs_score_unilateral'] is not None else '',
        show_answer_details=bool(session.get('is_admin')),
        has_answer_details=has_answer_details,
        right_answer_rows=build_answer_rows(questions_for_row, right_answers, row_lang),
        left_answer_rows=build_answer_rows(questions_for_row, left_answers, row_lang),
        unilateral_answer_rows=build_answer_rows(questions_for_row, unilateral_answers, row_lang),
        amputation_level_right_label=amputation_level_text(row['amputation_level_right']),
        amputation_level_left_label=amputation_level_text(row['amputation_level_left']),
    )


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    get_lang()
    if request.method == 'POST':
        pw = request.form.get('password', '')
        if pw and pw == ADMIN_PASSWORD:
            session['is_admin'] = True
            return redirect(url_for('admin_responses'))
        flash('Wrong password', 'error')
    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    session.pop('is_admin', None)
    return redirect(url_for('index'))


@app.route('/admin')
@admin_required
def admin_responses():
    db = get_db()
    rows = db.execute(f'SELECT * FROM responses WHERE {active_filter()} ORDER BY created_at DESC').fetchall()
    return render_template('admin.html', rows=rows)


@app.route('/admin/response/<int:response_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_response(response_id):
    get_lang()
    db = get_db()
    row = db.execute(f'SELECT * FROM responses WHERE id = ? AND {active_filter()}', (response_id,)).fetchone()
    if not row:
        abort(404)
    allowed_sides = ['bilateral'] if row['side'] == 'bilateral' else ['right', 'left']

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        age = parse_int_or_none(request.form.get('age'))
        language = request.form.get('language') if request.form.get('language') in LANGUAGES else row['language']
        procedure_type = request.form.get('procedure_type') if request.form.get('procedure_type') in PROCEDURES else row['procedure_type']
        side = request.form.get('side') if request.form.get('side') in allowed_sides else row['side']
        if not name:
            flash('Hasta adı boş bırakılamaz.', 'error')
            return redirect(url_for('admin_edit_response', response_id=response_id))

        sm_r = parse_int_or_none(request.form.get('surgery_month_right'))
        sy_r = parse_int_or_none(request.form.get('surgery_year_right'))
        sm_l = parse_int_or_none(request.form.get('surgery_month_left'))
        sy_l = parse_int_or_none(request.form.get('surgery_year_left'))
        amp_m_r = parse_int_or_none(request.form.get('amputation_month_right'))
        amp_y_r = parse_int_or_none(request.form.get('amputation_year_right'))
        amp_l_r = clean_text(request.form.get('amputation_level_right'))
        amp_m_l = parse_int_or_none(request.form.get('amputation_month_left'))
        amp_y_l = parse_int_or_none(request.form.get('amputation_year_left'))
        amp_l_l = clean_text(request.form.get('amputation_level_left'))

        if row['side'] in ('right', 'left') and side in ('right', 'left') and row['side'] != side:
            if side == 'left':
                sm_l = sm_l or sm_r or row['surgery_month_right']
                sy_l = sy_l or sy_r or row['surgery_year_right']
                amp_m_l = amp_m_l or amp_m_r or row['amputation_month_right']
                amp_y_l = amp_y_l or amp_y_r or row['amputation_year_right']
                amp_l_l = amp_l_l or amp_l_r or row['amputation_level_right']
            else:
                sm_r = sm_r or sm_l or row['surgery_month_left']
                sy_r = sy_r or sy_l or row['surgery_year_left']
                amp_m_r = amp_m_r or amp_m_l or row['amputation_month_left']
                amp_y_r = amp_y_r or amp_y_l or row['amputation_year_left']
                amp_l_r = amp_l_r or amp_l_l or row['amputation_level_left']

        is_amputee = procedure_type in AMPUTEE_PROCEDURES
        is_socket = procedure_type == 'socket_prosthesis'
        if procedure_type not in ('knee', 'hip', 'osseointegration'):
            sm_r = sy_r = sm_l = sy_l = None
        if procedure_type == 'osseointegration':
            amp_m_r = amp_m_l = None
        if not is_amputee:
            amp_m_r = amp_y_r = amp_l_r = None
            amp_m_l = amp_y_l = amp_l_l = None
        if side == 'right':
            sm_l = sy_l = amp_m_l = amp_y_l = amp_l_l = None
        elif side == 'left':
            sm_r = sy_r = amp_m_r = amp_y_r = amp_l_r = None

        fjs_score_right, fjs_score_left, fjs_score_unilateral = normalize_scores_for_side(row, side)
        db.execute(
            """UPDATE responses SET
               language = ?, name = ?, age = ?, procedure_type = ?, side = ?,
               surgery_month_right = ?, surgery_year_right = ?, surgery_month_left = ?, surgery_year_left = ?,
               amputation_month_right = ?, amputation_year_right = ?, amputation_level_right = ?,
               amputation_month_left = ?, amputation_year_left = ?, amputation_level_left = ?,
               fjs_score_right = ?, fjs_score_left = ?, fjs_score_unilateral = ?
               WHERE id = ?""",
            (
                language, name, age, procedure_type, side,
                sm_r, sy_r, sm_l, sy_l,
                amp_m_r, amp_y_r, amp_l_r,
                amp_m_l, amp_y_l, amp_l_l,
                fjs_score_right, fjs_score_left, fjs_score_unilateral,
                response_id,
            )
        )
        db.commit()
        flash('Hasta bilgileri güncellendi. Anket cevapları değiştirilmedi.', 'success')
        return redirect(url_for('result', response_id=response_id))

    current_year = datetime.now().year
    return render_template(
        'admin_edit_response.html',
        r=row,
        allowed_sides=allowed_sides,
        current_year=current_year,
        surgery_years=range(current_year, 1979, -1),
        amputation_years=range(current_year, 1949, -1),
        months=range(1, 13),
        languages=LANGUAGES,
        procedures=PROCEDURES,
        amputation_levels=AMPUTATION_LEVEL_LABELS,
    )


@app.route('/admin/response/<int:response_id>/archive', methods=['POST'])
@admin_required
def admin_archive_response(response_id):
    db = get_db()
    row = db.execute(f'SELECT id, name FROM responses WHERE id = ? AND {active_filter()}', (response_id,)).fetchone()
    if not row:
        flash('Kayıt bulunamadı.', 'error')
        return redirect(url_for('admin_responses'))
    db.execute("UPDATE responses SET is_archived = 1, archived_at = datetime('now') WHERE id = ?", (response_id,))
    db.commit()
    flash(f"{row['name']} kaydı listeden kaldırıldı.", 'success')
    return redirect(url_for('admin_responses'))


@app.route('/admin/export.csv')
@admin_required
def admin_export_csv():
    headers, data_rows = get_export_table()
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(data_rows)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Type': 'text/csv; charset=utf-8',
            'Content-Disposition': f'attachment; filename={export_filename("csv")}',
            'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0',
            'Pragma': 'no-cache',
            'Expires': '0',
        },
    )


@app.route('/admin/export.xlsx')
@admin_required
def admin_export_xlsx():
    headers, data_rows = get_export_table()
    workbook = ExcelWorkbook()
    sheet = workbook.active
    sheet.title = 'FJS Yanıtları'
    sheet.append(headers)
    for data_row in data_rows:
        sheet.append(data_row)

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(bottom=Side(style='thin', color='D9E2F3'))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        if col_letter in ('D', 'W'):
            width = min(max(max_len + 2, 18), 45)
        elif 'answer' in str(column_cells[0].value).lower() or col_letter == get_column_letter(len(headers)):
            width = min(max(max_len + 2, 18), 40)
        else:
            width = min(max(max_len + 2, 10), 24)
        sheet.column_dimensions[col_letter].width = width
    xlsx_io = io.BytesIO()
    workbook.save(xlsx_io)
    xlsx_io.seek(0)
    return Response(
        xlsx_io.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={export_filename("xlsx")}',
            'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0',
            'Pragma': 'no-cache',
            'Expires': '0',
        },
    )


@app.route('/healthz')
def healthz():
    return 'ok', 200


with app.app_context():
    init_db()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
